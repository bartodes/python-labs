import openpyxl


def songsPerPlaylist(working_sheet):
    playlist_songs = {} # Reading the amount of songs in a playlist

    for row in range(2, working_sheet.max_row + 1):
        playlist = working_sheet.cell(row, 4).value

        if playlist in playlist_songs:
            playlist_songs[playlist] += 1
        else:
            playlist_songs[playlist] = 1

    print(f"The amount of songs per playlist is: \n{playlist_songs}\n")


def listenersPerPlaylist(working_sheet):
    global playlist_listeners
    playlist_listeners = {} # Reading the total of listeners per playlist
    
    for row in range(2, working_sheet.max_row + 1):
        playlist = working_sheet.cell(row, 4).value
        listeners = working_sheet.cell(row, 2).value

        if playlist in playlist_listeners:
            playlist_listeners[playlist] += listeners
        else:
            playlist_listeners[playlist] = listeners

    print(f"The total of listeners of the songs in a playlist: \n{playlist_listeners}\n")


def listenersPerArtist(working_sheet):
    artist_listeners = {} # Reading the total of listeners per artist

    for row in range(2, working_sheet.max_row + 1):
        artist = working_sheet.cell(row, 3).value
        listeners = working_sheet.cell(row, 2).value
        
        if artist in artist_listeners:
            artist_listeners[artist] += listeners
        else:
            artist_listeners[artist] = listeners

    print(f"The total listeners per Artits: \n{artist_listeners}\n")


def playlistForgottenSongs(working_sheet):
    playlist_forgotten_songs = {} # Reading the amount of songs under 100,000 listeners in a playlist
    forgotten_value = 100000

    for row in range(2, working_sheet.max_row + 1):
        playlist = working_sheet.cell(row, 4).value
        listeners = working_sheet.cell(row, 2).value

        if listeners <= forgotten_value and playlist in playlist_forgotten_songs:
            playlist_forgotten_songs[playlist] += 1
        elif listeners <= forgotten_value:
            playlist_forgotten_songs[playlist] = 1
        
    print(f"The amount of songs under 100000 listeners: \n{playlist_forgotten_songs}\n")


def addResultsSheet(working_file):
    # Writing in a new sheet the total listeners from all the songs, that the playlist has.

    working_file.create_sheet("Playlist Popularity") 
    results_sheet = working_file["Playlist Popularity"]
    
    headers = ["N°","Playlists","Listeners"]
    for i in range(len(headers)):
        results_sheet.cell(1, i + 1).value = headers[i]

    values_list = list(playlist_listeners.values())
    values_list.sort(reverse=True)
    
    for i in range(2, len(playlist_listeners) + 2):
        results_sheet.cell(i, 1).value = i - 1
        results_sheet.cell(i, 3).value = values_list[i - 2]
        for k, v in playlist_listeners.items():
            if values_list[i - 2] == v:
                results_sheet.cell(i, 2).value = k

    working_file.save("songsFromPlaylist.xlsx")
    print("Results sheet named 'Playlist Popularity' has been added. \nCheck it out!")


def main():
    file = openpyxl.load_workbook("songsFromPlaylist.xlsx")
    sheet = file["Sheet1"]
    
    songsPerPlaylist(sheet)
    listenersPerPlaylist(sheet)
    listenersPerArtist(sheet)
    playlistForgottenSongs(sheet)
    addResultsSheet(file)
main()